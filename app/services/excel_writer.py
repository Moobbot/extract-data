"""
app/services/excel_writer.py

Ghi dữ liệu vào sheet Data của file Excel template.
Thay thế _save_combined_excel() khi template có excel_template_file.

Luồng:
  1. Copy file template → output path
  2. Mở workbook đã copy bằng openpyxl (giữ nguyên format/style)
  3. Tìm sheet Data (hoặc tên sheet cấu hình)
  4. Phát hiện dòng header (dòng đầu tiên có dữ liệu)
  5. Ghi dữ liệu từ dòng tiếp theo, căn theo cột header
  6. Lưu file
"""

from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


def _find_header_row(ws, sheet_name: str) -> tuple[int, list[str]]:
    """Tìm dòng header trong sheet.

    Cấu trúc template chuẩn:
      Row 1: Tiêu đề (merged cell) — ví dụ "DANH SÁCH VĂN BẰNG..."
      Row 2: Header cột (STT, Số hiệu bằng, ...)
      Row 3: Dòng mẫu bắt đầu bằng "Mẫu " (sẽ bị xóa khi ghi)

    Trả về (row_index_1based, [tên cột]).
    Header là dòng đầu tiên có ít nhất 2 ô không rỗng.
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:
            headers = [str(c).strip() if c is not None else "" for c in row]
            logger.debug(
                "Header tìm thấy tại row %d trong sheet '%s': %s",
                row_idx, sheet_name, headers[:5]
            )
            return row_idx, headers
    return 1, []


def _is_sample_row(ws, row_idx: int) -> bool:
    """Kiểm tra xem dòng có phải là dòng mẫu 'Mẫu ' trong template không."""
    row_values = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in ws[row_idx]
    ]
    first_nonempty = next((v for v in row_values if v), "")
    return first_nonempty.lower().startswith("mẫu")


def write_rows_to_template(
    rows: list[dict],
    template_path: str,
    output_path: str,
    sheet_name: str = "Data",
    start_stt: int = 1,
) -> Optional[str]:
    """Copy template và ghi rows vào sheet Data.

    Args:
        rows: Danh sách dict, mỗi dict là một hàng dữ liệu.
              Key là tên cột khớp với header trong template.
        template_path: Đường dẫn đến file Excel template gốc.
        output_path: Đường dẫn file Excel output sẽ tạo.
        sheet_name: Tên sheet để ghi dữ liệu (mặc định: "Data").
        start_stt: Giá trị STT bắt đầu (mặc định: 1).

    Returns:
        output_path nếu thành công, None nếu thất bại.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl chưa được cài. Chạy: pip install openpyxl")
        return None

    if not rows:
        logger.warning("write_rows_to_template: không có dữ liệu để ghi.")
        return None

    template_file = Path(template_path)
    if not template_file.exists():
        logger.error("File template không tồn tại: %s", template_path)
        return None

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Bước 1: Copy template → output (giữ nguyên format, công thức)
    shutil.copy2(template_file, output_file)
    logger.info(
        "Đã copy template '%s' → '%s'", template_file.name, output_file.name
    )

    try:
        wb = openpyxl.load_workbook(output_file)
    except Exception as e:
        logger.error("Không mở được file Excel: %s — %s", output_file, e)
        return None

    # Bước 2: Tìm sheet Data
    if sheet_name not in wb.sheetnames:
        logger.warning(
            "Sheet '%s' không tồn tại trong template. Các sheet: %s. "
            "Dùng sheet đầu tiên: '%s'.",
            sheet_name, wb.sheetnames, wb.sheetnames[0]
        )
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Bước 3: Phát hiện dòng header
    header_row_idx, headers = _find_header_row(ws, sheet_name)

    if not headers:
        logger.error(
            "Không tìm thấy header trong sheet '%s'. Ghi thất bại.", sheet_name
        )
        wb.close()
        return None

    # Tạo map: tên cột (lower stripped) → chỉ số cột (0-based)
    header_map: dict[str, int] = {}
    for col_idx, h in enumerate(headers):
        if h:
            header_map[h.strip().lower()] = col_idx

    # Bước 4: Xác định dòng bắt đầu dữ liệu
    #   Template chuẩn: Row 1 = tiêu đề, Row 2 = header, Row 3 = dòng mẫu "Mẫu"
    #   Dòng mẫu được GIỮ NGUYÊN, dữ liệu ghi từ row 4 trở đi.
    data_start_row = header_row_idx + 1

    # Nếu dòng ngay sau header là dòng mẫu → bỏ qua, ghi từ dòng tiếp theo
    if data_start_row <= ws.max_row and _is_sample_row(ws, data_start_row):
        logger.info(
            "Giữ nguyên dòng mẫu 'Mẫu' tại row %d — ghi dữ liệu từ row %d.",
            data_start_row, data_start_row + 1
        )
        data_start_row += 1  # Bỏ qua dòng mẫu

    # Xóa dữ liệu cũ từ data_start_row trở đi (không đụng dòng mẫu)
    max_row = ws.max_row
    if max_row >= data_start_row:
        for row in ws.iter_rows(min_row=data_start_row, max_row=max_row):
            for cell in row:
                cell.value = None

    # Bước 5: Ghi dữ liệu
    stt_keys = {"stt", "số thứ tự", "no", "no."}
    written = 0

    for row_offset, record in enumerate(rows):
        excel_row = data_start_row + row_offset

        for col_name, col_idx in header_map.items():
            cell = ws.cell(row=excel_row, column=col_idx + 1)

            # Tự động điền STT
            if col_name in stt_keys:
                cell.value = start_stt + row_offset
                continue

            # Tìm giá trị trong record (khớp không phân biệt hoa thường)
            value = None
            for record_key, record_val in record.items():
                if record_key.strip().lower() == col_name:
                    value = record_val
                    break

            if value is not None:
                cell.value = value if value != "" else None

        written += 1

    # Bước 6: Lưu
    try:
        wb.save(output_file)
        wb.close()
        logger.info(
            "Đã ghi %d dòng vào sheet '%s' của '%s'",
            written, sheet_name, output_file.name
        )
        return str(output_file)
    except Exception as e:
        logger.error("Lỗi khi lưu file Excel: %s — %s", output_file, e)
        try:
            wb.close()
        except Exception:
            pass
        return None


def _get_base_dir() -> str:
    """Trả về thư mục gốc project (chứa app/, config/, templates/)."""
    from app.core.config import settings
    return settings.BASE_DIR


def get_template_excel_path(template_id: str) -> Optional[str]:
    """Trả về đường dẫn tuyệt đối tới file Excel template theo template_id."""
    import json
    from app.core.config import settings

    config_path = os.path.join(settings.BASE_DIR, "config", "templates.json")
    templates_dir = settings.TEMPLATES_DIR

    logger.debug(
        "BASE_DIR=%s | templates_dir=%s | template_id=%s",
        settings.BASE_DIR, templates_dir, template_id
    )

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            templates = json.load(f)
    except Exception as e:
        logger.error("Không đọc được config/templates.json: %s", e)
        return None

    template_config = templates.get(template_id, {})
    filename = template_config.get("excel_template_file")
    if not filename:
        logger.warning(
            "template_id='%s' không có excel_template_file trong config.", template_id
        )
        return None

    full_path = os.path.join(templates_dir, filename)
    if not os.path.exists(full_path):
        logger.error(
            "File template không tìm thấy: %s\n"
            "Kiểm tra volume mount './templates:/app/templates:ro' trong docker-compose.yml.",
            full_path
        )
        return None

    logger.info("Dùng template Excel: %s", full_path)
    return full_path


def get_template_sheet_name(template_id: str) -> str:
    """Trả về tên sheet Data theo template_id (mặc định: 'Data')."""
    import json

    base_dir = _get_base_dir()
    config_path = os.path.join(base_dir, "config", "templates.json")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            templates = json.load(f)
        return templates.get(template_id, {}).get("excel_data_sheet", "Data")
    except Exception:
        return "Data"


