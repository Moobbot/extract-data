from openpyxl import Workbook, load_workbook

from app.services.excel_writer import write_rows_to_template


def test_write_rows_to_template_preserves_extracted_stt(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["DANH SÁCH"])
    ws.append(["STT", "Số hiệu bằng"])
    ws.append(["Mẫu ", "H129039"])
    wb.save(template)
    wb.close()

    saved = write_rows_to_template(
        rows=[
            {"STT": "16", "Số hiệu bằng": "506074"},
            {"STT": "17", "Số hiệu bằng": "506075"},
        ],
        template_path=str(template),
        output_path=str(output),
        sheet_name="Data",
    )

    assert saved == str(output)

    result = load_workbook(output, data_only=True)
    ws = result["Data"]
    assert ws["A4"].value == "16"
    assert ws["A5"].value == "17"
    assert ws["B4"].value == "506074"
    assert ws["B5"].value == "506075"
    result.close()
